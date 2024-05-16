#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import json

from openpyxl import load_workbook
import logging
import yaml
import uuid

class TreeNode(dict):  
    """
        初始化一个树节点对象实例。

        参数:
        id: 节点的唯一标识符。
        title: 节点的标题。
        level: 节点在树中的层级。
        parent_id: 父节点的标识符。
        path: 节点在树中的路径。

        属性:
        children: 一个空列表，用于存储该节点的子节点。
    """
    def __init__(self, id, title, level, parent_id, path, name, menu_type, ancestors, component, hideMenu, sort, children=None):
        super().__init__()
        self['id'] = id
        self['title'] = title
        self['level'] = level
        self['parent_id'] = parent_id
        self['path'] = path
        self['name'] = name
        self['menu_type'] = menu_type
        self['ancestors'] = ancestors
        self['component'] = component
        self['hideMenu'] = hideMenu
        self['sort'] = sort
        self['children'] = children or []

    def add_child(self, child):  
        self['children'].append(child)  
        
    def __json__(self):
       return {
            "id": self['id'],
            "title": self['title'],
            "level": self['level'],
            "parentId": self['parent_id'],
            'path': self['path'],
            'name': self['name'],
            'menu_type': self['menu_type'],
            'ancestors': self['ancestors'],
            'component': self['component'],
            'hideMenu': self['hideMenu'],
            'sort': self['sort'],
            # 列表推导式，遍历子节点，递归调用__json__方法
            "children": [child.__json__() for child in self['children']]
        }
    
class TreeDataExcelReader:
    """
       树形数据结构读取器
    """
    def __init__(self, logger):
        self.node_list = []
        self.root_node_list = []
        # 每一级的父级分开存放，方便循环中存取，最大级数6
        self.parent_node_list = [None] * 6

        self.row_num = 0
        self.logger = logger
        pass

    # 根据excel中数据构建树结构
    def buildTreeDataFromExcel(self, excel_config_definitions):
        workbook = load_workbook(excel_config_definitions['excel_file_path'])
        sheet_name = "菜单结构"
        # 查看sheet页清单
        # workbook.sheet_names()
        # print("sheets：" + str(workBook.sheet_names()))
        sheet = workbook[sheet_name]

        # 遍历sheet页中的行
        for row in sheet.iter_rows(min_row=excel_config_definitions['data_row'], values_only=True):
            level = 0
            name = ''
            node = TreeNode(None, None, None, None, None, None, None, None, None, None, None)
            for colNum in range(excel_config_definitions['name_start'], excel_config_definitions['name_end']):
                name = row[colNum]
                if name :
                    node['id'] = excel_config_definitions['generate_start_id'] + self.row_num
                    node['title'] = name
                    node['level'] = level + 1
                    if level > 0:
                        parent_node = self.parent_node_list[level-1]
                        if parent_node:
                            parent_node.add_child(node)
                            node['parent_id'] = parent_node['id']
                            # 菜单祖辈路径字段
                            node['ancestors'] = f'{parent_node["ancestors"]},{parent_node["id"]}'
                    else:
                        node['ancestors'] = '0'
                        node['parent_id'] = '0'
                        self.root_node_list.append(node)
                    self.parent_node_list[level] = node
                    break
                else:
                    level = level +1
            # 菜单属性字段
            col_mapping = excel_config_definitions['col_mapping']
            for colName, colNum in col_mapping.items():
                if colName:
                    node[colName] = row[colNum]
            if node['path'] is None:
                node['path'] = node['id']
            if node['menu_type'] is None or node['menu_type'] == "":
                node['menu_type'] = "M"
            if node['menu_type'] == 'C' and (node['name'] is None or node['name'] == ""):
                node['name'] = uuid.uuid4().__str__() 
            if node['title']:
                self.node_list.append(node)
                self.logger.info(node)
            self.row_num = self.row_num + 1
        



if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    excelReader = TreeDataExcelReader(logging)
    with open('config.yaml', 'r') as file:
        config = yaml.safe_load(file)
    excel_config = config['excel_config']
    excelReader.buildTreeDataFromExcel(excel_config)
    print(json.dumps([child.__json__() for child in excelReader.root_node_list], ensure_ascii=False, indent=4))
    print(excelReader.row_num)