import json
import logging
import psycopg2
import yaml
from openpyxl import load_workbook

from tree_node import TreeNode



# 创建一个日志记录器实例
logger = logging.getLogger(__name__)
# 根据需要设置日志级别
logger.setLevel(logging.DEBUG)  


class TreeDataExcelReader:
    """
       树形数据结构读取器
    """
    def __init__(self, logger):
        self.node_list = []
        self.root_node_list = []
        # 每一级的父级分开存放，方便循环中存取，最大级数6
        self.parent_node_list = [None] * 6

        self.row_num = 0
        self.logger = logger
        pass

    # 根据excel中数据构建树结构
    def buildTreeDataFromExcel(self, excel_config_definitions):
        workbook = load_workbook(excel_config_definitions['excel_file_path'])
        sheet_name = excel_config_definitions['sheet_name']
        # 查看sheet页清单
        # workbook.sheet_names()
        # print("sheets：" + str(workBook.sheet_names()))
        sheet = workbook[sheet_name]

        # 遍历sheet页中的行
        for row in sheet.iter_rows(min_row=excel_config_definitions['data_row'], values_only=True):
            level = 0
            name = ''
            node = TreeNode(None, None, None, None, None, None, None, None, None, None, None)
            for colNum in range(excel_config_definitions['name_start'], excel_config_definitions['name_end']):
                name = row[colNum]
                if name is not None and name.strip() != '' :
                    node['id'] = excel_config_definitions['generate_start_id'] + self.row_num
                    node['title'] = name.strip()
                    node['level'] = level + 1
                    if level > 0:
                        parent_node = self.parent_node_list[level-1]
                        if parent_node:
                            parent_node.add_child(node)
                            node['parent_id'] = parent_node['id']
                            # 菜单祖辈路径字段
                            node['ancestors'] = f'{parent_node["ancestors"]},{parent_node["id"]}'
                    else:
                        node['ancestors'] = '0'
                        node['parent_id'] = '0'
                        self.root_node_list.append(node)
                    self.parent_node_list[level] = node
                    break
                else:
                    level = level +1
            # 菜单属性字段
            col_mapping = excel_config_definitions['col_mapping']
            for colName, colNum in col_mapping.items():
                if colName:
                    node[colName] = row[colNum]
            if node['path'] is None:
                node['path'] = node['id']
            if node['menu_type'] is None or node['menu_type'] == "":
                node['menu_type'] = "M"
            if node['menu_type'] == 'C' and (node['name'] is None or node['name'] == ""):
                node['name'] = uuid.uuid4().__str__() 
            if node['title']:
                self.node_list.append(node)
                self.logger.info(node)
            self.row_num = self.row_num + 1
        

# 导入数据到 Kingbase 数据库
def importDataToKingbaseDB(dbconfig, table_name, menu_list):
    schema_name = dbconfig['schema_name']
    # 插入语句
    sql = f'''
    INSERT INTO "{schema_name}"."{table_name}" (id,parent_id,name,title,menu_type, delevel, ancestors,"path",frame_src,component,
    param_path,transition_name,ignore_route,is_cache,is_affix,is_disabled,frame_type,hide_tab,hide_menu,hide_breadcrumb,
    hide_children,hide_path_for_children,dynamic_level,real_path,perms,icon,sort,status,remark,create_by,create_time,
    update_by,update_time,is_common,is_default,del_flag,module_id,tenant_id) 
    VALUES
    (%(id)s,%(parent_id)s,%(name)s,%(title)s,%(menu_type)s,%(level)s,%(ancestors)s,%(path)s,NULL,%(component)s,NULL,NULL,'N','Y','N','N','0','0',%(hideMenu)s,'0','0','0',5,NULL,'',NULL,%(sort)s,'0',
    '',NULL,'2023-11-15 14:20:04',NULL,NULL,'1','Y',0,1,0);

    '''

    with psycopg2.connect(
        host=dbconfig['host'],
        port=dbconfig['port'],
        user=dbconfig['user'],
        password=dbconfig['password'],
        database=dbconfig['dbname'],
    ) as conn:
        with conn.cursor() as cursor:
            # 使用executemany()方法，传入SQL和字典列表
            logging.info(menu_list)
            try:
                cursor.executemany(sql, menu_list)
                conn.commit()
            except Exception as e:
                logging.error("执行executemany失败", exc_info=True)


if __name__ == '__main__':

    with open('config.yaml', 'r', encoding='utf-8') as file:
        config = yaml.safe_load(file)
    excel_config = config['excel_config']
    dbconfig = config['database']
  
    excelReader = TreeDataExcelReader(logger)

    # 读取excel文件
    excelReader.buildTreeDataFromExcel(excel_config)

    # 插入到数据库
    importDataToKingbaseDB(dbconfig, excel_config['table_name'], excelReader.node_list)